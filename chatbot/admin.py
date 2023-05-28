import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from django.contrib import admin
from .models import ChatHistory

class ChatHistoryAdmin(admin.ModelAdmin):
    list_display = ['user', 'prompt', 'message']
    actions = ['export_to_excel']  # Menambahkan aksi export_to_excel

    def export_to_excel(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="chat_history.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Chat History"

        columns = ['User', 'Prompt', 'Message']

        # Menulis header kolom
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f"{column_letter}1"] = column_title

        # Menulis data
        for row_num, chat in enumerate(queryset, 2):
            worksheet[f"A{row_num}"] = chat.user.username
            worksheet[f"B{row_num}"] = chat.prompt
            worksheet[f"C{row_num}"] = chat.message

        # Mengatur lebar kolom
        for col_num in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 15

        workbook.save(response)
        return response

    export_to_excel.short_description = "Export to Excel"

admin.site.register(ChatHistory, ChatHistoryAdmin)
