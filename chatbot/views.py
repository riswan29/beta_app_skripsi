from django.shortcuts import render, redirect
import requests
import openai
from openpyxl import Workbook,load_workbook
from django.contrib.auth import authenticate, login as django_login
from django.contrib.auth.decorators import login_required
from .models import *
from django.core.paginator import Paginator
from django.db.models import Q
from .key import Keys
import datetime
from django.http import HttpResponse
from app.models import UserProfile, Jadwal

openai.api_key = Keys

# Daftar pesan dalam sesi
session_messages = []
# Daftar pencarian sebelumnya
search_history = []
MAX_HISTORY_LENGTH = 20  # Jumlah maksimum riwayat pencarian yang ditampilkan
SHORTENED_LENGTH = 20  # Panjang maksimum riwayat yang dipersingkat
ITEMS_PER_PAGE = 5  # Jumlah item yang ditampilkan per halaman

@login_required(login_url="login")
def homeBot(request):
    if request.method == "POST":
        prompt = request.POST.get("prompt")
        model_engine = "text-davinci-003"

        if prompt.startswith("/jadwaldosen"):
            # Ambil hari saat ini
            today = datetime.datetime.now().strftime("%A")
            # Ambil jadwal dosen untuk hari ini dari model Jadwal
            jadwal_hari_ini = Jadwal.objects.filter(hari=today)

            # Buat pesan balasan dengan jadwal semua dosen
            response = "Jadwal dosen hari ini:\n"
            if jadwal_hari_ini.exists():
                for jadwal in jadwal_hari_ini:
                    response += f"Dosen: {jadwal.dosen.full_name}\n"
                    response += f"Hari: {jadwal.hari}\n"
                    response += f"Jam: {jadwal.waktu} - {jadwal.waktu_selesai}\n"
                    response += f"Mata Kuliah: {jadwal.nama_mata_kuliah}\n"
                    response += f"Ruangan: {jadwal.ruangan}\n"
                    response += "\n"
            else:
                response = "Tidak ada jadwal dosen hari ini."

            chat_history = ChatHistory(user=request.user, prompt=prompt, message=response)
            chat_history.save()

            session_messages.append({"sender": "user", "content": prompt})
            session_messages.append({"sender": "bot", "content": response})
        else:
            completions = openai.Completion.create(
                engine=model_engine,
                prompt=prompt,
                max_tokens=1024,
                n=1,
                temperature=0.5,
            )
            response = completions.choices[0].text

            chat_history = ChatHistory(user=request.user, prompt=prompt, message=response)
            chat_history.save()

            session_messages.append({"sender": "user", "content": prompt})
            session_messages.append({"sender": "bot", "content": response})

    search_query = request.GET.get('search_query')

    # Display the user's search history
    search_history = ChatHistory.objects.filter(user=request.user).order_by("-id")

    if search_query:
        search_history = search_history.filter(Q(prompt__icontains=search_query) | Q(message__icontains=search_query))

    paginator = Paginator(search_history, ITEMS_PER_PAGE)
    page_number = request.GET.get("page", 1)
    page = paginator.get_page(page_number)

    # Limit jumlah history yang ditampilkan sesuai halaman yang dipilih
    start_index = (page.number - 1) * ITEMS_PER_PAGE
    end_index = start_index + ITEMS_PER_PAGE
    search_history = page[start_index:end_index]

    context = {
        "messages": session_messages,
        "searches": search_history,
        "page": page,
        "search_query": search_query,
    }

    return render(request, "indexx.html", context)


@login_required(login_url="login")
def newChat(request):
    global session_messages
    session_messages = []
    return render(request, "indexx.html", {})

@login_required(login_url="login")
def loadChat(request, search_id):
    try:
        # Mengambil respons yang tersimpan dalam database berdasarkan ID pencarian
        chat_history = ChatHistory.objects.get(id=search_id)
        prompt = chat_history.prompt
        message = chat_history.message

        context = {"prompt": prompt, "message": message}
        return render(request, "histo.html", context)
    except ChatHistory.DoesNotExist:
        # Tidak ada respons yang tersimpan dalam database untuk ID pencarian yang dipilih
        return redirect("chatbot")

@login_required(login_url="login")
def save_to_excel(prompt, message):
    # Buka file Excel atau buat file baru jika belum ada
    try:
        wb = load_workbook('chatlog.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1).value = 'Prompt'
        sheet.cell(row=1, column=2).value = 'Message'

    # Cari baris terakhir yang sudah terisi
    last_row = sheet.max_row + 1

    # Tambahkan data baru ke baris terakhir
    sheet.cell(row=last_row, column=1).value = prompt
    sheet.cell(row=last_row, column=2).value = message

    # Simpan file Excel
    wb.save('chatlog.xlsx')
